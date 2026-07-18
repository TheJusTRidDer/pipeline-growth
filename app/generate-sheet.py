from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
DARK = "1a1a2e"
BLUE = "2563eb"
GRAY = "8a8aaa"
LIGHT = "f1f5f9"
GREEN = "22c55e"
ORANGE = "f59e0b"
RED = "ef4444"
PURPLE = "8b5cf6"

hdr_font = Font(name="Inter", bold=True, color="ffffff", size=11)
hdr_fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
sub_font = Font(name="Inter", bold=True, color=DARK, size=10)
body_font = Font(name="Inter", color=DARK, size=10)
title_font = Font(name="Inter", bold=True, color=DARK, size=16)
section_font = Font(name="Inter", bold=True, color=BLUE, size=12)
thin_border = Border(
    left=Side(style="thin", color="e2e8f0"),
    right=Side(style="thin", color="e2e8f0"),
    top=Side(style="thin", color="e2e8f0"),
    bottom=Side(style="thin", color="e2e8f0"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

stage_colors = {
    "Prospect": ("e0f2fe", "0369a1"),
    "Contacted": ("fef3c7", "b45309"),
    "Responded": ("fef3c7", "b45309"),
    "Discovery Call": ("ede9fe", "6d28d9"),
    "Audit in Progress": ("ede9fe", "6d28d9"),
    "Proposal Sent": ("fce7f3", "be185d"),
    "Negotiation": ("fce7f3", "be185d"),
    "Closed Won": ("dcfce7", "15803d"),
    "Closed Lost": ("fee2e2", "991b1b"),
    "Nurture": ("f3e8ff", "7e22ce"),
}

# ===== SHEET 1: Lead Tracker =====
ws1 = wb.active
ws1.title = "Lead Tracker"
ws1.sheet_properties.tabColor = BLUE

headers = [
    "Company", "Website", "Contact Name", "Title", "Email", "LinkedIn",
    "Industry", "Employees", "Revenue", "Location", "Priority Score",
    "Stage", "Pipeline Value ($)", "Lead Source", "Email Step",
    "Created Date", "Notes"
]
col_widths = [18, 22, 16, 18, 28, 28, 18, 10, 12, 16, 12, 16, 14, 12, 12, 14, 30]

# Title
ws1.merge_cells("A1:Q1")
ws1["A1"] = "Pipeline Growth — Lead Tracker"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:Q2")
ws1["A2"] = "Copy data from the CRM app or enter manually. Priority Score 1-10. Stages match the pipeline."
ws1["A2"].font = Font(name="Inter", color=GRAY, size=9, italic=True)
ws1.row_dimensions[2].height = 20

# Headers row 4
for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=4, column=col_idx, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = thin_border
    ws1.column_dimensions[get_column_letter(col_idx)].width = w

ws1.row_dimensions[4].height = 28

# Data rows (50 empty rows with validation hints)
for r in range(5, 55):
    for c in range(1, len(headers) + 1):
        cell = ws1.cell(row=r, column=c)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = left_wrap
    ws1.cell(row=r, column=12).value = "Prospect"
    ws1.cell(row=r, column=11).value = 5

# Stage dropdown hint
ws1.merge_cells("A55:Q55")
ws1["A55"] = "Stages: Prospect | Contacted | Responded | Discovery Call | Audit in Progress | Proposal Sent | Negotiation | Closed Won | Closed Lost | Nurture"
ws1["A55"].font = Font(name="Inter", color=GRAY, size=9, italic=True)

# Conditional formatting notes
ws1.merge_cells("A56:Q56")
ws1["A56"] = "TIP: In Excel, use Data > Data Validation > List for Stage column (L) with the stage names above."
ws1["A56"].font = Font(name="Inter", color=GRAY, size=9, italic=True)

# ===== SHEET 2: Pipeline Summary =====
ws2 = wb.create_sheet("Pipeline Summary")
ws2.sheet_properties.tabColor = GREEN

ws2.merge_cells("A1:H1")
ws2["A1"] = "Pipeline Summary Dashboard"
ws2["A1"].font = title_font
ws2.row_dimensions[1].height = 36

# Metrics section
ws2["A3"] = "KEY METRICS"
ws2["A3"].font = section_font

metrics = [
    ("Total Leads", "=COUNTA('Lead Tracker'!C5:C54)", BLUE),
    ("Active Pipeline", '=COUNTIF(\'Lead Tracker\'!L5:L54,"Prospect")+COUNTIF(\'Lead Tracker\'!L5:L54,"Contacted")+COUNTIF(\'Lead Tracker\'!L5:L54,"Responded")+COUNTIF(\'Lead Tracker\'!L5:L54,"Discovery Call")+COUNTIF(\'Lead Tracker\'!L5:L54,"Audit in Progress")+COUNTIF(\'Lead Tracker\'!L5:L54,"Proposal Sent")+COUNTIF(\'Lead Tracker\'!L5:L54,"Negotiation")', ORANGE),
    ("Closed Won", '=COUNTIF(\'Lead Tracker\'!L5:L54,"Closed Won")', GREEN),
    ("Closed Lost", '=COUNTIF(\'Lead Tracker\'!L5:L54,"Closed Lost")', RED),
    ("Pipeline Value ($)", '=SUMPRODUCT((\'Lead Tracker\'!L5:L54<>"Closed Won")*(\'Lead Tracker\'!L5:L54<>"Closed Lost")*(\'Lead Tracker\'!L5:L54<>"Nurture")*(\'Lead Tracker\'!M5:M54))', PURPLE),
    ("Won Revenue ($)", '=SUMIF(\'Lead Tracker\'!L5:L54,"Closed Won",\'Lead Tracker\'!M5:M54)', GREEN),
    ("Avg Deal Size ($)", '=IFERROR(SUMIF(\'Lead Tracker\'!L5:L54,"Closed Won",\'Lead Tracker\'!M5:M54)/COUNTIF(\'Lead Tracker\'!L5:L54,"Closed Won"),0)', BLUE),
    ("Conversion Rate", '=IFERROR(COUNTIF(\'Lead Tracker\'!L5:L54,"Closed Won")/COUNTA(\'Lead Tracker\'!L5:L54),0)', GREEN),
]

for i, (label, formula, color) in enumerate(metrics):
    row = 4 + i
    cell_l = ws2.cell(row=row, column=1, value=label)
    cell_l.font = Font(name="Inter", bold=True, color=DARK, size=10)
    cell_l.border = thin_border
    cell_l.alignment = left_wrap
    cell_v = ws2.cell(row=row, column=2, value=formula if "=" in str(formula) else formula)
    cell_v.font = Font(name="Inter", bold=True, color=color, size=14)
    cell_v.border = thin_border
    cell_v.alignment = center
    if "%" in label:
        cell_v.number_format = "0%"
    elif "$" in label:
        cell_v.number_format = '$#,##0_);($#,##0)'
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18

# Stage breakdown
ws2["A13"] = "PIPELINE BY STAGE"
ws2["A13"].font = section_font

stage_headers = ["Stage", "Count", "Value ($)", "Avg Score"]
for c, h in enumerate(stage_headers, 1):
    cell = ws2.cell(row=14, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = thin_border

stages = ["Prospect", "Contacted", "Responded", "Discovery Call", "Audit in Progress",
          "Proposal Sent", "Negotiation", "Closed Won", "Closed Lost", "Nurture"]

for i, stage in enumerate(stages):
    row = 15 + i
    ws2.cell(row=row, column=1, value=stage).font = body_font
    ws2.cell(row=row, column=1).border = thin_border
    ws2.cell(row=row, column=1).alignment = left_wrap
    bg, fg = stage_colors.get(stage, ("fff", DARK))
    ws2.cell(row=row, column=1).fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    ws2.cell(row=row, column=1).font = Font(name="Inter", bold=True, color=fg, size=10)

    ws2.cell(row=row, column=2, value=f"=COUNTIF('Lead Tracker'!L$5:L$54,\"{stage}\")").font = body_font
    ws2.cell(row=row, column=2).border = thin_border
    ws2.cell(row=row, column=2).alignment = center

    ws2.cell(row=row, column=3, value=f"=SUMIF('Lead Tracker'!L$5:L$54,\"{stage}\",'Lead Tracker'!M$5:M$54)").font = body_font
    ws2.cell(row=row, column=3).border = thin_border
    ws2.cell(row=row, column=3).alignment = center
    ws2.cell(row=row, column=3).number_format = '$#,##0_);($#,##0)'

    ws2.cell(row=row, column=4, value=f"=IFERROR(AVERAGEIF('Lead Tracker'!L$5:L$54,\"{stage}\",'Lead Tracker'!K$5:K$54),0)").font = body_font
    ws2.cell(row=row, column=4).border = thin_border
    ws2.cell(row=row, column=4).alignment = center
    ws2.cell(row=row, column=4).number_format = "0.0"

# Totals
row_total = 15 + len(stages)
ws2.cell(row=row_total, column=1, value="TOTAL").font = Font(name="Inter", bold=True, color=DARK, size=10)
ws2.cell(row=row_total, column=1).border = thin_border
ws2.cell(row=row_total, column=2, value=f"=SUM(B15:B{row_total-1})").font = Font(name="Inter", bold=True, size=10)
ws2.cell(row=row_total, column=2).border = thin_border
ws2.cell(row=row_total, column=2).alignment = center
ws2.cell(row=row_total, column=3, value=f"=SUM(C15:C{row_total-1})").font = Font(name="Inter", bold=True, size=10)
ws2.cell(row=row_total, column=3).border = thin_border
ws2.cell(row=row_total, column=3).alignment = center
ws2.cell(row=row_total, column=3).number_format = '$#,##0_);($#,##0)'
ws2.cell(row=row_total, column=4, value=f"=AVERAGE(D15:D{row_total-1})").font = Font(name="Inter", bold=True, size=10)
ws2.cell(row=row_total, column=4).border = thin_border
ws2.cell(row=row_total, column=4).alignment = center
ws2.cell(row=row_total, column=4).number_format = "0.0"

# ===== SHEET 3: Email Sequences =====
ws3 = wb.create_sheet("Email Sequences")
ws3.sheet_properties.tabColor = ORANGE

ws3.merge_cells("A1:G1")
ws3["A1"] = "Email Sequence Tracker"
ws3["A1"].font = title_font
ws3.row_dimensions[1].height = 36

email_headers = ["Company", "Contact", "Email 1 Sent", "Email 2 Sent", "Email 3 Sent", "Email 4 Sent", "Reply Received?"]
for c, h in enumerate(email_headers, 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = thin_border

ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 18
for col in range(3, 8):
    ws3.column_dimensions[get_column_letter(col)].width = 16

for r in range(4, 54):
    for c in range(1, 8):
        cell = ws3.cell(row=r, column=c)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = center
    ws3.cell(row=r, column=3).value = "☐"
    ws3.cell(row=r, column=4).value = "☐"
    ws3.cell(row=r, column=5).value = "☐"
    ws3.cell(row=r, column=6).value = "☐"
    ws3.cell(row=r, column=7).value = "☐"

ws3.cell(row=54, column=1, value='☐ = Not sent | ☑ = Sent | Enter reply date in column G').font = Font(name="Inter", color=GRAY, size=9, italic=True)

# ===== SHEET 4: Weekly Targets =====
ws4 = wb.create_sheet("Weekly Targets")
ws4.sheet_properties.tabColor = PURPLE

ws4.merge_cells("A1:E1")
ws4["A1"] = "Weekly Performance Tracker"
ws4["A1"].font = title_font
ws4.row_dimensions[1].height = 36

ws4.merge_cells("A2:E2")
ws4["A2"] = "Track weekly against targets from the CRM Pipeline Setup Guide"
ws4["A2"].font = Font(name="Inter", color=GRAY, size=9, italic=True)

tracker_headers = ["Week Starting", "New Prospects Added", "Emails Sent", "Replies Received", "Discovery Calls Booked"]
targets = [50, 200, 5, 2]
for c, h in enumerate(tracker_headers, 1):
    cell = ws4.cell(row=4, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = thin_border

ws4.column_dimensions["A"].width = 18
for col in range(2, 6):
    ws4.column_dimensions[get_column_letter(col)].width = 22

for r in range(5, 25):
    for c in range(1, 6):
        cell = ws4.cell(row=r, column=c)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = center

week_start = 5
for r in range(5, 25):
    ws4.cell(row=r, column=1)

# Target row
ws4.cell(row=26, column=1, value="TARGET").font = Font(name="Inter", bold=True, color=DARK, size=10)
ws4.cell(row=26, column=1).border = thin_border
target_texts = ["", 50, 200, 5, 2]
for c, v in enumerate(target_texts, 2):
    cell = ws4.cell(row=26, column=c, value=v)
    cell.font = Font(name="Inter", bold=True, color=BLUE, size=10)
    cell.border = thin_border
    cell.alignment = center

# Average row
ws4.cell(row=27, column=1, value="AVERAGE").font = Font(name="Inter", bold=True, color=DARK, size=10)
ws4.cell(row=27, column=1).border = thin_border
for c in range(2, 6):
    col = get_column_letter(c)
    cell = ws4.cell(row=27, column=c, value=f"=AVERAGE({col}5:{col}24)")
    cell.font = Font(name="Inter", bold=True, color=PURPLE, size=10)
    cell.border = thin_border
    cell.alignment = center
    cell.number_format = "0.0"

# ===== SHEET 5: Monthly Revenue =====
ws5 = wb.create_sheet("Monthly Revenue")
ws5.sheet_properties.tabColor = GREEN

ws5.merge_cells("A1:C1")
ws5["A1"] = "Monthly Revenue Tracking"
ws5["A1"].font = title_font
ws5.row_dimensions[1].height = 36

rev_headers = ["Month", "New Clients", "Revenue ($)"]
for c, h in enumerate(rev_headers, 1):
    cell = ws5.cell(row=3, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = thin_border

ws5.column_dimensions["A"].width = 18
ws5.column_dimensions["B"].width = 14
ws5.column_dimensions["C"].width = 18

months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
for i, m in enumerate(months):
    row = 4 + i
    ws5.cell(row=row, column=1, value=m).font = body_font
    ws5.cell(row=row, column=1).border = thin_border
    ws5.cell(row=row, column=1).alignment = center
    ws5.cell(row=row, column=2, value=0).font = body_font
    ws5.cell(row=row, column=2).border = thin_border
    ws5.cell(row=row, column=2).alignment = center
    ws5.cell(row=row, column=3, value=0).font = body_font
    ws5.cell(row=row, column=3).border = thin_border
    ws5.cell(row=row, column=3).alignment = center
    ws5.cell(row=row, column=3).number_format = '$#,##0_);($#,##0)'

row_total = 4 + len(months)
ws5.cell(row=row_total, column=1, value="TOTAL").font = Font(name="Inter", bold=True, size=10)
ws5.cell(row=row_total, column=1).border = thin_border
ws5.cell(row=row_total, column=1).alignment = center
ws5.cell(row=row_total, column=2, value=f"=SUM(B4:B{row_total-1})").font = Font(name="Inter", bold=True, size=10)
ws5.cell(row=row_total, column=2).border = thin_border
ws5.cell(row=row_total, column=2).alignment = center
ws5.cell(row=row_total, column=3, value=f"=SUM(C4:C{row_total-1})").font = Font(name="Inter", bold=True, size=10)
ws5.cell(row=row_total, column=3).border = thin_border
ws5.cell(row=row_total, column=3).alignment = center
ws5.cell(row=row_total, column=3).number_format = '$#,##0_);($#,##0)'

# Print settings
for ws in [ws1, ws2, ws3, ws4, ws5]:
    ws.sheet_view.showGridLines = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

output = "Pipeline-Growth-Spreadsheet.xlsx"
wb.save(output)
print(f"Created {output}")
